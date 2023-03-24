from tkinter import *
import random
import openpyxl
from ttkbootstrap import Style
from tkinter import Label
from tkinter import ttk
import threading
import time
import ttkbootstrap as ttk
from time import sleep

person=['李明阳','刘宗翰','张子涵','王一伊','赵天麒','高芷韵',
'周健军','马浩文','田浩然','程诗洋','曹起豪','陈美希','王玉策',
'杨佳星','刘峰宇','殷  兴','王一婷','赵健凯','费  硕','张美娆',
'牛嘉熙','邵枳博','易思涵','王欣然','于  启','闵  睿','边诩诗',
'王艺霖','于知画','左松洋','刘则谦','丁  敏','徐诗涛','田书宇',
'王禹鑫','何佳滢','张峰瑞','白雪松','赵宏哲','佟香凝','李兆星',
'吴思淼','周  硕','孙祥涵','陈佳馨','智兰淇','刘超冉','张馨心',
'杨连鑫']
yin=['窗1','窗2','窗3','中南','中中','中北','门1','门2','门3']
number=list(range(5,51,5))
filename='rizhi.txt'
huiche='\n'

stop_sign=False
stop_id=None
going=True
is_run=False

def get_time():     #左上角时间函数
    time2=time.strftime('%Y-%m-%d %H:%M:%S')
    clock=Label(window,text=time2,font=28)
    clock.place(x=0,y=0)
    clock.after(1000,get_time)
def com1(event):      #下拉菜单中学生姓名绑定学号
    quzhi=com.get()
    lab["text"]=quzhi
    if quzhi=='李明阳':
        lab2["text"]='01'
    elif quzhi=='刘宗翰':
        lab2['text']='02'
    elif quzhi=='张子涵':
        lab2['text']='03'
    elif quzhi=='王一伊':
        lab2['text']='04'
    elif quzhi=='赵天麒':
        lab2['text']='05'
    elif quzhi=='高芷韵':
        lab2['text']='06'
    elif quzhi=='周健军':
        lab2['text']='07'
    elif quzhi=='马浩文':
        lab2['text']='08'
    elif quzhi=='田浩然':
        lab2['text']='09'
    elif quzhi=='程诗洋':
        lab2['text']='10'
    elif quzhi=='曹起豪':
        lab2['text']='11'
    elif quzhi=='陈美希':
        lab2['text']='12'
    elif quzhi=='王玉策':
        lab2['text']='13'
    elif quzhi=='杨佳星':
        lab2['text']='14'
    elif quzhi=='刘峰宇':
        lab2['text']='16'
    elif quzhi=='殷  兴':
        lab2['text']='17'
    elif quzhi=='王一婷':
        lab2['text']='18'
    elif quzhi=='赵健凯':
        lab2['text']='19'
    elif quzhi=='费  硕':
        lab2['text']='20'
    elif quzhi=='张美娆':
        lab2['text']='21'
    elif quzhi=='牛嘉熙':
        lab2['text']='22'
    elif quzhi=='邵枳博':
        lab2['text']='23'
    elif quzhi=='易思涵':
        lab2['text']='24'
    elif quzhi=='王欣然':
        lab2['text']='25'
    elif quzhi=='于  启':
        lab2['text']='26'
    elif quzhi=='闵  睿':
        lab2['text']='27'
    elif quzhi=='边诩诗':
        lab2['text']='28'
    elif quzhi=='王艺霖':
        lab2['text']='29'
    elif quzhi=='于知画':
        lab2['text']='30'
    elif quzhi=='左松洋':
        lab2['text']='31'
    elif quzhi=='刘则谦':
        lab2['text']='32'
    elif quzhi=='丁  敏':
        lab2['text']='33'
    elif quzhi=='徐诗涛':
        lab2['text']='34'
    elif quzhi=='田书宇':
        lab2['text']='35'
    elif quzhi=='王禹鑫':
        lab2['text']='36'
    elif quzhi=='何佳滢':
        lab2['text']='38'
    elif quzhi=='张峰瑞':
        lab2['text']='39'
    elif quzhi=='白雪松':
        lab2['text']='40'
    elif quzhi=='赵宏哲':
        lab2['text']='41'
    elif quzhi=='佟香凝':
        lab2['text']='42'
    elif quzhi=='李兆星':
        lab2['text']='43'
    elif quzhi=='吴思淼':
        lab2['text']='44'
    elif quzhi=='周  硕':
        lab2['text']='46'
    elif quzhi=='孙祥涵':
        lab2['text']='47'
    elif quzhi=='陈佳馨':
        lab2['text']='48'
    elif quzhi=='智兰淇':
        lab2['text']='49'
    elif quzhi=='刘超冉':
        lab2['text']='50'
    elif quzhi=='张馨心':
        lab2['text']='51'
    elif quzhi=='杨连鑫':
        lab2['text']='00'
def click_button1():
    with open(filename,'a') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        a=int(sr1.get())
        b=int(sr2.get())
        c=random.randint(a,b)
        txt.insert(END,c)
        txt.insert(END,huiche)
        file_object.write("[" + str(a) + "," + str(b) + "]" + str(c) + "\n")
def click_button2():
    with open(filename,'a') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        d=random.choice(person)
        txt.insert(END,d)
        txt.insert(END,huiche)
        file_object.write("随机选了: " + d + "\n")
def click_button3():
    with open(filename,'a') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        k=int(sr3.get())
        e=random.sample(person,k)
        txt.insert(END,e)
        txt.insert(END,huiche)
        file_object.write("从四班随机选了: " + str(k) + "个人_分别是: " + str(e) + "\n")
def click_button4():
    with open(filename,'a') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        f=random.choice(number)
        txt.insert(END,f)
        txt.insert(END,huiche)
        file_object.write('随机选了: ' + str(f) + '\n')
def click_button5():
    with open(filename,'a') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        g=random.choice(yin)
        txt.insert(END,g)
        txt.insert(END,huiche)
        file_object.write('阿谦的活: ' + g+'\n')
def click_button6():
    with open(filename,'a') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        data=openpyxl.load_workbook('excel_test.xlsx')
        sheetnames=data.get_sheet_names()
        ws=data.get_sheet_by_name(sheetnames[0])
        ws=data.active
        bh=int(sr4.get())
        xyz=random.sample(person,bh)
        txt.insert(END,xyz)
        txt.insert(END,huiche)
        asd=int(sr5.get())
        file_object.write("从四班随机选了: " + str(bh) + "个人_分别是: " + str(xyz) + "\n")
        if '刘则谦' in xyz:
            ws.cell(row=2,column=int(asd),value='√')
        if '邵枳博' in xyz:
            ws.cell(row=3,column=int(asd),value='√')
        if '赵天麒' in xyz:
            ws.cell(row=4,column=int(asd),value='√')
        if '于启' in xyz:
            ws.cell(row=5,column=int(asd),value='√')
        if '白雪松' in xyz:
            ws.cell(row=6,column=int(asd),value='√')
        if '张美娆' in xyz:
            ws.cell(row=7,column=int(asd),value='√')
        if '王禹鑫' in xyz:
            ws.cell(row=8,column=int(asd),value='√')
        if '王玉策' in xyz:
            ws.cell(row=9,column=int(asd),value='√')
        if '杨佳星' in xyz:
            ws.cell(row=10,column=int(asd),value='√')
        if '高芷韵' in xyz:
            ws.cell(row=11,column=int(asd),value='√')
        if '刘超冉' in xyz:
            ws.cell(row=12,column=int(asd),value='√')
        if '智兰淇' in xyz:
            ws.cell(row=13,column=int(asd),value='√')
        if '周健军' in xyz:
            ws.cell(row=14,column=int(asd),value='√')
        if '孙祥涵' in xyz:
            ws.cell(row=15,column=int(asd),value='√')
        if '程诗洋' in xyz:
            ws.cell(row=16,column=int(asd),value='√')
        if '费硕' in xyz:
            ws.cell(row=17,column=int(asd),value='√')
        if '周硕' in xyz:
            ws.cell(row=18,column=int(asd),value='√')
        if '王欣然' in xyz:
            ws.cell(row=19,column=int(asd),value='√')
        if '陈佳馨' in xyz:
            ws.cell(row=20,column=int(asd),value='√')
        if '赵宏哲' in xyz:
            ws.cell(row=21,column=int(asd),value='√')
        if '殷兴' in xyz:
            ws.cell(row=22,column=int(asd),value='√')
        if '王一伊' in xyz:
            ws.cell(row=23,column=int(asd),value='√')
        if '田书宇' in xyz:
            ws.cell(row=24,column=int(asd),value='√')
        if '赵健凯' in xyz:
            ws.cell(row=25,column=int(asd),value='√')
        if '王艺霖' in xyz:
            ws.cell(row=26,column=int(asd),value='√')
        if '刘宗翰' in xyz:
            ws.cell(row=27,column=int(asd),value='√')
        if '李明阳' in xyz:
            ws.cell(row=28,column=int(asd),value='√')
        if '左松洋' in xyz:
            ws.cell(row=29,column=int(asd),value='√')
        if '马浩文' in xyz:
            ws.cell(row=30,column=int(asd),value='√')
        if '边诩诗' in xyz:
            ws.cell(row=31,column=int(asd),value='√')
        if '易思涵' in xyz:
            ws.cell(row=32,column=int(asd),value='√')
        if '陈美希' in xyz:
            ws.cell(row=33,column=int(asd),value='√')
        if '李兆星' in xyz:
            ws.cell(row=34,column=int(asd),value='√')
        if '刘峰宇' in xyz:
            ws.cell(row=35,column=int(asd),value='√')
        if '张峰瑞' in xyz:
            ws.cell(row=36,column=int(asd),value='√')
        if '闵睿' in xyz:
            ws.cell(row=37,column=int(asd),value='√')
        if '何佳滢' in xyz:
            ws.cell(row=38,column=int(asd),value='√')
        if '丁敏' in xyz:
            ws.cell(row=39,column=int(asd),value='√')
        if '佟香凝' in xyz:
            ws.cell(row=40,column=int(asd),value='√')
        if '曹起豪' in xyz:
            ws.cell(row=41,column=int(asd),value='√')
        if '田浩然' in xyz:
            ws.cell(row=42,column=int(asd),value='√')
        if '徐诗涛' in xyz:
            ws.cell(row=43,column=int(asd),value='√')
        if '张馨心' in xyz:
            ws.cell(row=44,column=int(asd),value='√')
        if '张子涵' in xyz:
            ws.cell(row=45,column=int(asd),value='√')
        if '杨连鑫' in xyz:
            ws.cell(row=46,column=int(asd),value='√')
        if '王一婷' in xyz:
            ws.cell(row=47,column=int(asd),value='√')
        if '牛嘉熙' in xyz:
            ws.cell(row=48,column=int(asd),value='√')
        if '于知画' in xyz:
            ws.cell(row=49,column=int(asd),value='√')
        if '吴思淼' in xyz:
            ws.cell(row=50,column=int(asd),value='√')
        data.save('excel_test.xlsx')
def lottery_roll(var1,var2):
    global going
    show_member=random.choice(person)
    var1.set(show_member)
    if going:
        window.after(50,lottery_roll,var1,var2)
    else:
        var2.set('选中 {} ！！！'.format(show_member))
        going=True
        with open(filename,'a') as file_object:
            now=time.strftime('%Y-%m-%d %H:%M:%S')
            file_object.write(now)
            file_object.write(show_member)
        return
def lottery_start(var1,var2):
    global is_run
    if is_run:
        return
    is_run=True
    var2.set('按 结 束 停 止')
    lottery_roll(var1,var2)
def lottery_end():
    global going,is_run
    if is_run:
        going=False
        is_run=False
def round():   # 定义一个函数： 1）循环选项；2）改变选项的颜色
    global stop_id
    i=1
    if isinstance(stop_id,int):
        i=stop_id
    while True:
        time.sleep(0.008)
        for x in button_list:
            x['bg']='#375A7F'
        button_list[i]['bg']='#00BC8C'
        i += 1
        if i>= len(button_list):
            i=0
        if stop_sign == True:
            stop_id=i
            with open(filename,'a') as file_object:
                now=time.strftime('%Y-%m-%d %H:%M:%S')
                file_object.write(now)
                file_object.write('\n'+str(stop_id)+'\n')
            break
def stop():    # 定义停止的方法
    global stop_sign
    if stop_sign == True:
        return
    stop_sign=True
def start():   # 定义开始的方法 （开启线程，单独跑一个循环的函数）
    global stop_sign
    stop_sign=False
    t=threading.Thread(target=round)
    t.start() # 创建线程并启动

window=ttk.Window(title="班级抽签系统_v1.1",themename="darkly",size=(1090,690),resizable=None,alpha=0.9)   #创建tkinter窗口

#击鼓传花学生姓名
btn1=Button(window,text='李明阳',bg='white')
btn1.place(x=600,y=0,width=50,height=50)
btn2=Button(window,text='刘宗翰',bg='white')
btn2.place(x=600,y=60,width=50,height=50)
btn3=Button(window,text='张子涵',bg='white')
btn3.place(x=600,y=120,width=50,height=50)
btn4=Button(window,text='王一伊',bg='white')
btn4.place(x=600,y=180,width=50,height=50)
btn5=Button(window,text='赵天麒',bg='white')
btn5.place(x=600,y=240,width=50,height=50)
btn6=Button(window,text='高芷韵',bg='white')
btn6.place(x=600,y=300,width=50,height=50)
btn7=Button(window,text='周健军',bg='white')
btn7.place(x=600,y=360,width=50,height=50)
btn8=Button(window,text='马浩文',bg='white')
btn8.place(x=600,y=420,width=50,height=50)
btn9=Button(window,text='田浩然',bg='white')
btn9.place(x=660,y=420,width=50,height=50)
btn10=Button(window,text='程诗洋',bg='white')
btn10.place(x=720,y=420,width=50,height=50)
btn11=Button(window,text='曹起豪',bg='white')
btn11.place(x=780,y=420,width=50,height=50)
btn12=Button(window,text='陈美希',bg='white')
btn12.place(x=840,y=420,width=50,height=50)
btn13=Button(window,text='王玉策',bg='white')
btn13.place(x=900,y=420,width=50,height=50)
btn14=Button(window,text='杨佳星',bg='white')
btn14.place(x=960,y=420,width=50,height=50)
btn15=Button(window,text='刘峰宇',bg='white')
btn15.place(x=1020,y=420,width=50,height=50)
btn16=Button(window,text='殷  兴',bg='white')
btn16.place(x=1020,y=360,width=50,height=50)
btn17=Button(window,text='王一婷',bg='white')
btn17.place(x=1020,y=300,width=50,height=50)
btn18=Button(window,text='赵健凯',bg='white')
btn18.place(x=1020,y=240,width=50,height=50)
btn19=Button(window,text='费  硕',bg='white')
btn19.place(x=1020,y=180,width=50,height=50)
btn20=Button(window,text='张美娆',bg='white')
btn20.place(x=1020,y=120,width=50,height=50)
btn21=Button(window,text='牛嘉熙',bg='white')
btn21.place(x=1020,y=60,width=50,height=50)
btn22=Button(window,text='邵枳博',bg='white')
btn22.place(x=1020,y=0,width=50,height=50)
btn23=Button(window,text='易思涵',bg='white')
btn23.place(x=960,y=0,width=50,height=50)
btn24=Button(window,text='王欣然',bg='white')
btn24.place(x=900,y=0,width=50,height=50)
btn25=Button(window,text='于  启',bg='white')
btn25.place(x=840,y=0,width=50,height=50)
btn26=Button(window,text='闵  睿',bg='white')
btn26.place(x=780,y=0,width=50,height=50)
btn27=Button(window,text='边诩诗',bg='white')
btn27.place(x=720,y=0,width=50,height=50)
btn28=Button(window,text='王艺霖',bg='white')
btn28.place(x=660,y=0,width=50,height=50)
btn29=Button(window,text='于知画',bg='white')
btn29.place(x=660,y=60,width=50,height=50)
btn30=Button(window,text='左松洋',bg='white')
btn30.place(x=660,y=120,width=50,height=50)
btn31=Button(window,text='刘则谦',bg='white')
btn31.place(x=660,y=180,width=50,height=50)
btn32=Button(window,text='丁  敏',bg='white')
btn32.place(x=660,y=240,width=50,height=50)
btn33=Button(window,text='徐诗涛',bg='white')
btn33.place(x=660,y=300,width=50,height=50)
btn34=Button(window,text='田书宇',bg='white')
btn34.place(x=660,y=360,width=50,height=50)
btn35=Button(window,text='王禹鑫',bg='white')
btn35.place(x=720,y=360,width=50,height=50)
btn36=Button(window,text='何佳滢',bg='white')
btn36.place(x=780,y=360,width=50,height=50)
btn37=Button(window,text='张峰瑞',bg='white')
btn37.place(x=840,y=360,width=50,height=50)
btn38=Button(window,text='白雪松',bg='white')
btn38.place(x=900,y=360,width=50,height=50)
btn39=Button(window,text='赵宏哲',bg='white')
btn39.place(x=960,y=360,width=50,height=50)
btn40=Button(window,text='佟香凝',bg='white')
btn40.place(x=960,y=300,width=50,height=50)
btn41=Button(window,text='李兆星',bg='white')
btn41.place(x=960,y=240,width=50,height=50)
btn42=Button(window,text='吴思淼',bg='white')
btn42.place(x=960,y=180,width=50,height=50)
btn43=Button(window,text='周  硕',bg='white')
btn43.place(x=960,y=120,width=50,height=50)
btn44=Button(window,text='孙祥涵',bg='white')
btn44.place(x=960,y=60,width=50,height=50)
btn45=Button(window,text='陈佳馨',bg='white')
btn45.place(x=900,y=60,width=50,height=50)
btn46=Button(window,text='智兰淇',bg='white')
btn46.place(x=840,y=60,width=50,height=50)
btn47=Button(window,text='刘超冉',bg='white')
btn47.place(x=780,y=60,width=50,height=50)
btn48=Button(window,text='张馨心',bg='white')
btn48.place(x=720,y=60,width=50,height=50)
btn49=Button(window,text='杨连鑫',bg='white')
btn49.place(x=720,y=120,width=50,height=50)

button_list=[btn1,btn2,btn3,btn4,btn5,btn6,btn7,btn8,btn9,btn10,btn11,btn12,btn13,btn14,btn15,btn16,btn17,btn18,btn19,btn20,btn21,btn22,
btn23,btn24,btn25,btn26,btn27,btn28,btn29,btn30,btn31,btn32,btn33,btn34,btn35,btn36,btn37,btn38,btn39,btn40,btn41,btn42,btn43,btn44,btn45,
btn46,btn47,btn48,btn49]

btn_start=ttk.Button(window,text='开始',style='success.Outline.TButton',command=start)
btn_start.place(x=780,y=210,width=50,height=50)
btn_stop=ttk.Button(window,text='停止',style='success.Outline.TButton',command=stop)
btn_stop.place(x=840,y=210,width=50,height=50)

zhanwei=Label(window,text=" ")
zhanwei.grid(row=0,column=0,sticky=W)
bn1=ttk.Button(window,text='姓  名',style='success.Outline.TButton')
bn1.grid(row=1,column=0,sticky=W)
lab=Label(window,text="喜羊羊")
lab.grid(row=1,column=1,sticky=W)
bn2=ttk.Button(window,text='学  号',style='success.Outline.TButton')
bn2.grid(row=1,column=2,sticky=W)
lab2=Label(window,text="00")
lab2.grid(row=1,column=3,sticky=W)
com=ttk.Combobox(window,value=('李明阳','刘宗翰','张子涵','王一伊','赵天麒','高芷韵',
'周健军','马浩文','田浩然','程诗洋','曹起豪','陈美希','王玉策',
'杨佳星','刘峰宇','殷  兴','王一婷','赵健凯','费  硕','张美娆',
'牛嘉熙','邵枳博','易思涵','王欣然','于  启','闵  睿','边诩诗',
'王艺霖','于知画','左松洋','刘则谦','丁  敏','徐诗涛','田书宇',
'王禹鑫','何佳滢','张峰瑞','白雪松','赵宏哲','佟香凝','李兆星',
'吴思淼','周  硕','孙祥涵','陈佳馨','智兰淇','刘超冉','张馨心',
'杨连鑫'),bootstyle='success')     # #创建下拉菜单
com.grid(row=1,column=4,sticky=W)     # #将下拉菜单绑定到窗体
sr1=ttk.Entry(window,bootstyle='success')
sr1.place(x=60,y=100)
sr2=ttk.Entry(window,bootstyle='success')
sr2.place(x=220,y=100)
sr3=ttk.Entry(window,bootstyle='success')
sr3.place(x=60,y=200)
sr4=ttk.Entry(window,bootstyle='success')
sr4.place(x=60,y=350)
sr5=ttk.Entry(window,bootstyle='success')
sr5.place(x=220,y=350)
bn3=ttk.Button(window,text='执  行',style='success.Outline.TButton',command=click_button1)
bn3.place(x=0,y=100)
bn4=ttk.Button(window,text='执  行',style='success.Outline.TButton',command=click_button2)
bn4.place(x=0,y=150)
bn5=ttk.Button(window,text='执  行',style='success.Outline.TButton',command=click_button3)
bn5.place(x=0,y=200)
bn6=ttk.Button(window,text='执  行',style='success.Outline.TButton',command=click_button4)
bn6.place(x=0,y=250)
bn7=ttk.Button(window,text='执  行',style='success.Outline.TButton',command=click_button5)
bn7.place(x=0,y=300)
bn8=ttk.Button(window,text='执  行',style='success.Outline.TButton',command=click_button6)
bn8.place(x=0,y=350)
bq1=ttk.Button(window,text='随机生成一个[a,b]范围内的整数',style='success.Outline.TButton',command=click_button6)
bq1.place(x=380,y=100)
bq2=ttk.Button(window,text='从四班随机挑选一个人',style='success.Outline.TButton',command=click_button6)
bq2.place(x=380,y=150)
bq3=ttk.Button(window,text='从四班随机挑选k个人',style='success.Outline.TButton',command=click_button6)
bq3.place(x=380,y=200)
bq4=ttk.Button(window,text='随机挑选一个在<=50被5整除的数',style='success.Outline.TButton',command=click_button6)
bq4.place(x=380,y=250)
bq5=ttk.Button(window,text='随机行列',style='success.Outline.TButton',command=click_button6)
bq5.place(x=380,y=300)
bq6=ttk.Button(window,text='从四班随机挑选z个人{并载入excel}',style='success.Outline.TButton',command=click_button6)
bq6.place(x=380,y=350)

com.current(0)    # #设定下拉菜单的默认值为第1个
com.bind("<<ComboboxSelected>>",com1)
get_time()
txt=Text(window)
txt.place(rely=0.6,relheight=0.4)
if __name__ == '__main__': 
    var1=StringVar(value='即 将 开 始')
    show_label1=Label(window,textvariable=var1,justify='left',anchor=CENTER,width=17,height=3,font='楷体 -40 bold')
    show_label1.place(x=654,y=480)
    var2=StringVar(value='即 将 开 始')
    show_label2=Label(window,textvariable=var2,justify='left',anchor=CENTER,width=38,height=3,font='楷体 -18 bold')
    show_label2.place(x=650,y=600)
    button1=ttk.Button(window,text='开  始',style='success.Outline.TButton',command=lambda: lottery_start(var1,var2))
    button1.place(x=980,y=500)
    button2=ttk.Button(window,text='结  束',style='success.Outline.TButton',command=lambda: lottery_end())
    button2.place(x=980,y=640)
window.mainloop()