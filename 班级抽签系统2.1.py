import ttkbootstrap as ttk
import random
import threading
from tkinter import *
import time
import openpyxl
import ttkbootstrap as ttk

person=['李明阳','刘宗翰','张子涵','王一伊','赵天麒','高芷韵','周健军','马浩文','田浩然','程诗洋','曹起豪','陈美希','王玉策',
'杨佳星','刘峰宇','殷  兴','王一婷','赵健凯','费  硕','张美娆','牛嘉熙','邵枳博','易思涵','王欣然','于  启','闵  睿','边诩诗',
'王艺霖','于知画','左松洋','刘则谦','丁  敏','徐诗涛','田书宇','王禹鑫','何佳滢','张峰瑞','白雪松','赵宏哲','佟香凝','李兆星',
'吴思淼','周  硕','孙祥涵','陈佳馨','智兰淇','刘超冉','张馨心','杨连鑫']
yin=['窗1','窗2','窗3','中南','中中','中北','门1','门2','门3']
number=list(range(5,51,5))
filename= '../../rizhi.txt'
huiche='\n'
welcome1='\ \      / /__| | ___ ___  _ __ ___   ___  \n'
welcome2=' \ \ /\ / / _ \ |/ __/ _ \|  _   _ \ / _ \ \n'
welcome3='  \ V  V /  __/ | (_| (_) | | | | | |    __/ \n'
welcome4='   \_/\_/ \___|_|\___\___/|_| |_| |_|\___| \n'
welcome=welcome1+welcome2+welcome3+welcome4
stop_sign=False
stop_id=None
going=True
is_run=False

def click_button1():
    with open(filename,'a') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        a=int(sr1.get())
        b=int(sr2.get())
        c=random.randint(a,b)
        txt.insert(END,str(c))
        txt.insert(END,huiche)
        txt.see(END)
        file_object.write("["+str(a)+","+str(b)+"]"+str(c)+"\n")
def click_button2():
    with open(filename,'a') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        d=random.choice(person)
        txt.insert(END,d)
        txt.insert(END,huiche)
        txt.see(END)
        file_object.write("随机选了: "+d+"\n")
def click_button3():
    with open(filename,'a') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        k=int(sr3.get())
        e=random.sample(person,k)
        txt.insert(END,e)
        txt.insert(END,huiche)
        txt.see(END)
        file_object.write("从四班随机选了: "+str(k)+"个人_分别是: "+str(e)+"\n")
def click_button4():
    with open(filename,'a') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        f=random.choice(number)
        txt.insert(END,f)
        txt.insert(END,huiche)
        txt.see(END)
        file_object.write('随机选了: '+str(f)+'\n')
def click_button5():
    with open(filename,'a') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        g=random.choice(yin)
        txt.insert(END,g)
        txt.insert(END,huiche)
        txt.see(END)
        file_object.write('阿谦的活: '+g+'\n')
def click_button6():
    with open(filename,'a') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        data=openpyxl.load_workbook('excel_test.xlsx')
        sheetnames=data.get_sheet_names()
        ws=data.get_sheet_by_name(sheetnames[0])
        ws=data.active
        bh=int(sr4.get())
        xyz=random.sample(person,bh)
        txt.insert(END,xyz)
        txt.insert(END,huiche)
        txt.see(END)
        asd=int(sr5.get())
        file_object.write("从四班随机选了: "+str(bh)+"个人_分别是: "+str(xyz)+"\n")
        for rrrr,coco in zip(person,range(2,51)):
            if rrrr in xyz:
                ws.cell(row=coco,column=int(asd),value='√')
        nowxlsx=time.strftime('%Y-%m-%d')
        ws.cell(row=1,column=int(asd),value=nowxlsx)
        data.save('excel_test.xlsx')
def lottery_roll(var1,var2):
    global going
    show_member=random.choice(person)
    var1.set(show_member)
    if going:
        window.after(50,lottery_roll,var1,var2)
    else:
        var2.set('选中 {} ！！！'.format(show_member))
        going=True
        with open(filename,'a') as file_object:
            now=time.strftime('%Y-%m-%d %H:%M:%S')
            file_object.write(now)
            file_object.write(show_member+'\n')
        return
def lottery_start(var1,var2):
    global is_run
    if is_run:
        return
    is_run=True
    var2.set('按 结 束 停 止')
    lottery_roll(var1,var2)
def lottery_end():
    global going,is_run
    if is_run:
        going=False
        is_run=False
def round():   # 定义一个函数： 1）循环选项；2）改变选项的颜色
    global stop_id
    i=1
    if isinstance(stop_id,int):
        i=stop_id
    while True:
        time.sleep(0.008)
        for x in btn_list:
            x['bg']='#5289E9'
        btn_list[i]['bg']='#0C0C0C'
        i += 1
        if i>=len(btn_list):
            i=0
        if stop_sign==True:
            stop_id=i
            with open(filename,'a') as file_object:
                now=time.strftime('%Y-%m-%d %H:%M:%S')
                file_object.write(now)
                file_object.write('\t'+str(stop_id)+'\n')
            break
def stop():    # 定义停止的方法
    global stop_sign
    if stop_sign==True:
        return
    stop_sign=True
def start():   # 定义开始的方法 （开启线程，单独跑一个循环的函数）
    global stop_sign
    stop_sign=False
    t=threading.Thread(target=round)
    t.start() # 创建线程并启动
def qk():
    txt.delete('1.0',END)
window=ttk.Window(title="班级抽签系统_v2.1",themename="抽签",size=(1090,690),resizable=None,alpha=0.9)   #创建tkinter窗口

btn_start=ttk.Button(window,text='开始',style='success.Outline.TButton',command=start)
btn_start.place(x=750,y=210,width=50,height=50)
btn_stop=ttk.Button(window,text='停止',style='success.Outline.TButton',command=stop)
btn_stop.place(x=870,y=210,width=50,height=50)

kuang1 = ttk.Labelframe(window, text="随机生成一个在[a,b]范围内的整数", padding=10,bootstyle='success')
kuang1.pack(side=TOP,anchor=W,pady=5,padx=5)
kuang2 = ttk.Labelframe(window, text="从四班学生列表中随机挑选 一个人", padding=10,bootstyle='success')
kuang2.pack(side=TOP,anchor=W,pady=5,padx=5)
kuang3 = ttk.Labelframe(window, text="从四班学生列表中随机挑选 k 个人", padding=5,bootstyle='success')
kuang3.pack(side=TOP,anchor=W,pady=5,padx=5)
kuang4 = ttk.Labelframe(window, text="随机挑选一个在<=50被5整除的数", padding=10,bootstyle='success')
kuang4.pack(side=TOP,anchor=W,pady=5,padx=5)
kuang5 = ttk.Labelframe(window, text="随  机  挑  选  一 行  或  者  一 列", padding=10,bootstyle='success')
kuang5.pack(side=TOP,anchor=W,pady=5,padx=5)
kuang6 = ttk.Labelframe(window, text="从四班随机挑选z个人,并载入excel", padding=10,bootstyle='success')
kuang6.pack(side=TOP,anchor=W,pady=5,padx=5)

sr1=ttk.Entry(master=kuang1,bootstyle='success')
sr1.pack()
sr2=ttk.Entry(master=kuang1,bootstyle='success')
sr2.pack(pady=10)
sr3=ttk.Entry(master=kuang3,bootstyle='success')
sr3.pack(pady=5)
sr4=ttk.Entry(master=kuang6,bootstyle='success')
sr4.pack()
sr5=ttk.Entry(master=kuang6,bootstyle='success')
sr5.pack(pady=10)
bn3=ttk.Button(master=kuang1,text='执  行',style='success.Outline.TButton',command=click_button1)
bn3.pack()
bn4=ttk.Button(master=kuang2,text='执  行',style='success.Outline.TButton',command=click_button2)
bn4.pack()
bn5=ttk.Button(master=kuang3,text='执  行',style='success.Outline.TButton',command=click_button3)
bn5.pack(pady=5)
bn6=ttk.Button(master=kuang4,text='执  行',style='success.Outline.TButton',command=click_button4)
bn6.pack()
bn7=ttk.Button(master=kuang5,text='执  行',style='success.Outline.TButton',command=click_button5)
bn7.pack()
bn8=ttk.Button(master=kuang6,text='执  行',style='success.Outline.TButton',command=click_button6)
bn8.pack()

txt=ttk.ScrolledText(window, height=39, width=50)
txt.place(x=210,y=1)

var1=StringVar(value='即 将 开 始')
show_label1=Label(window,textvariable=var1,justify='left',anchor=CENTER,width=17,height=3,font='楷体 -40 bold')
show_label1.place(x=654,y=480)
var2=StringVar(value='即 将 开 始')
show_label2=Label(window,textvariable=var2,justify='left',anchor=CENTER,width=38,height=3,font='楷体 -18 bold')
show_label2.place(x=650,y=600)
button1=ttk.Button(window,text='开  始',style='success.Outline.TButton',command=lambda: lottery_start(var1,var2))
button1.place(x=980,y=500)
button2=ttk.Button(window,text='结  束',style='success.Outline.TButton',command=lambda: lottery_end())
button2.place(x=980,y=640)
qk=ttk.Button(window,text='清  空',style='success.Outline.TButton',command=qk)
qk.place(x=510,y=635)
txt.insert(END,welcome)
txt.insert(END,huiche)

place_x=['600','600','600','600','600','600','600','600','660','720','780','840','900','960','1020','1020','1020','1020','1020','1020','1020','1020','960','900','840','780','720','660','660','660','660','660','660','660','720','780','840','900','960','960','960','960','960','960','900','840','780','720','720']
place_y=['0','60','120','180','240','300','360','420','420','420','420','420','420','420','420','360','300','240','180','120','60','0','0','0','0','0','0','0','60','120','180','240','300','360','360','360','360','360','360','300','240','180','120','60','60','60','60','60','120']
btn_list=[]
for aaaa,bbbb,cccc in zip(place_x,place_y,person):
    button=Button(window,text=cccc,bg='white')
    button.place(x=aaaa,y=bbbb,width=50,height=50)
    btn_list.append(button)

style = ttk.Style()
theme_names = style.theme_names()
theme_selected = ttk.Label(window, text="抽签", font="-size 24 -weight bold",bootstyle='success')
theme_selected.place(x=1075,y=1)
theme_cbo = ttk.Combobox(window,text=style.theme.name,values=theme_names,bootstyle='success')
theme_cbo.place(x=1075,y=60)
theme_cbo.current(theme_names.index(style.theme.name))
def change_theme(e):
    t = theme_cbo.get()
    style.theme_use(t)
    theme_selected.configure(text=t)
    theme_cbo.selection_clear()
theme_cbo.bind("<<ComboboxSelected>>", change_theme)

window.mainloop()

